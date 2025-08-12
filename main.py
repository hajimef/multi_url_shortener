import sys
import os
import configparser
import requests
import openpyxl
from PyQt6.QtWidgets import QApplication, QMainWindow, QFileDialog, QDialog
from PyQt6.uic import loadUi
from PyQt6.QtCore import QThread, pyqtSignal

class ShortenWorker(QThread):
    progress_updated = pyqtSignal(int, str)
    log_updated = pyqtSignal(str)
    finished = pyqtSignal(str)

    def __init__(self, filepath, api_key):
        super().__init__()
        self.filepath = filepath
        self.api_key = api_key
        self.is_running = True

    def run(self):
        try:
            workbook = openpyxl.load_workbook(self.filepath)
            sheet = workbook.active
            
            urls_to_shorten = []
            for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    if cell.value:
                        urls_to_shorten.append(cell)

            total_urls = len(urls_to_shorten)
            if total_urls == 0:
                self.finished.emit("短縮するURLが見つかりませんでした。")
                return

            for i, cell in enumerate(urls_to_shorten):
                if not self.is_running:
                    break

                url = cell.value
                self.progress_updated.emit(int((i / total_urls) * 100), url)
                
                try:
                    api_url = f"https://xgd.io/V1/shorten?url={url}&key={self.api_key}"
                    response = requests.get(api_url)
                    response_json = response.json()

                    if response_json.get("status") == 200:
                        short_url = response_json.get("shorturl")
                        sheet.cell(row=cell.row, column=3).value = short_url
                        self.log_updated.emit(f"成功: {url} -> {short_url}")
                    else:
                        error_message = response_json.get("error", "不明なエラー")
                        self.log_updated.emit(f"失敗: {url} ({error_message})")

                except requests.RequestException as e:
                    self.log_updated.emit(f"APIリクエストエラー: {url} ({e})")
                except Exception as e:
                    self.log_updated.emit(f"エラー: {url} ({e})")

            self.progress_updated.emit(100, "完了")
            
            if self.is_running:
                save_path = self.filepath
                workbook.save(save_path)
                self.finished.emit(f"処理が完了しました。ファイルは {save_path} に保存されました。")
            else:
                self.finished.emit("処理が中断されました。")

        except FileNotFoundError:
            self.finished.emit("エラー: ファイルが見つかりません。")
        except Exception as e:
            self.finished.emit(f"予期せぬエラーが発生しました: {e}")

    def stop(self):
        self.is_running = False

class SettingsDialog(QDialog):
    def __init__(self, config_path):
        super().__init__()
        loadUi(os.path.join(os.path.dirname(os.path.abspath(__file__)), "dialog.ui"), self)
        self.config_path = config_path
        self.config = configparser.ConfigParser()
        self.load_settings()

    def load_settings(self):
        if os.path.exists(self.config_path):
            self.config.read(self.config_path)
            api_key = self.config.get("Settings", "api_key", fallback="")
            self.lineEdit_apiKey.setText(api_key)

    def accept(self):
        self.config["Settings"] = {"api_key": self.lineEdit_apiKey.text()}
        with open(self.config_path, "w") as configfile:
            self.config.write(configfile)
        super().accept()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi(os.path.join(os.path.dirname(os.path.abspath(__file__)), "interface.ui"), self)

        # setting.iniをスクリプトと同じディレクトリに保存
        self.config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "setting.ini")
        self.worker = None

        self.pushButton_selectFile.clicked.connect(self.select_file)
        self.pushButton_start.clicked.connect(self.start_shortening)
        self.pushButton_settings.clicked.connect(self.open_settings)
        
        self.pushButton_start.setText("短縮開始")

    def select_file(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "Excelファイルを選択", "", "Excel Files (*.xlsx)")
        if filepath:
            self.lineEdit_filepath.setText(filepath)
            self.textEdit_log.clear()

    def open_settings(self):
        dialog = SettingsDialog(self.config_path)
        dialog.exec()

    def start_shortening(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.pushButton_start.setText("短縮開始")
            return

        filepath = self.lineEdit_filepath.text()
        if not filepath:
            self.textEdit_log.setText("ファイルを選択してください。")
            return

        config = configparser.ConfigParser()
        if not os.path.exists(self.config_path) or not config.read(self.config_path) or not config.has_option("Settings", "api_key"):
            self.textEdit_log.setText("APIキーが設定されていません。[設定]からAPIキーを入力してください。")
            return
        
        api_key = config.get("Settings", "api_key")
        if not api_key:
            self.textEdit_log.setText("APIキーが空です。[設定]からAPIキーを入力してください。")
            return

        self.textEdit_log.clear()
        self.pushButton_start.setText("処理中... (クリックして中断)")
        self.worker = ShortenWorker(filepath, api_key)
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.log_updated.connect(self.update_log)
        self.worker.finished.connect(self.on_finished)
        self.worker.start()

    def update_progress(self, value, url):
        self.progressBar.setValue(value)
        self.label_current_url.setText(url)

    def update_log(self, message):
        self.textEdit_log.append(message)

    def on_finished(self, message):
        self.textEdit_log.append(message)
        self.progressBar.setValue(100)
        self.pushButton_start.setText("短縮開始")
        self.worker = None

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait()
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
